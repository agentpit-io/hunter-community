#!/usr/bin/env python3
"""生成港股主表 CSV —— `_24` §8.2⑤。

## 为什么要这个脚本

`hk.master`(港股主表)原来是「平台自建」的三条之一:代码 → 中文名 /
每手股数,走我们的网关或我们的数据库。开源用户拿不到,而它其实
**只是一张静态对照表** —— 不是服务、不需要 key、不需要实时。

所以导成 CSV 放进仓库,用户开箱就有。这条从"平台依赖"降级成"仓库里的
一个文件"。

## 数据来自港交所官方

`https://www.hkex.com.hk/.../ListOfSecurities.xlsx` —— 港交所每个交易日
更新的证券列表,权威且免费。字段:

    Stock Code | Name of Securities | Category | Sub-Category | Board Lot | ISIN

我们只取前五个里有用的四个。**不取 ISIN** —— 用不上,而且它会让 CSV
大一圈。

## 名称是英文

港交所这张表给的是英文名(`CKH HOLDINGS`),不是中文名。原来那张
`hk_stock_master` 有中文名和繁体名 —— 那是我们自己补的,不在公开数据里。

**这里不假装有中文名。**宁可只给英文名,也不要从别处凑一份对不上的中文
翻译:代码-名称对错一个,用户看到的就是另一家公司。缺中文名的影响只是
显示不好看,而错的中文名会让人下错单。

## 用法

    python scripts/gen_hk_master_csv.py            # 写到 data/hk_master.csv
    python scripts/gen_hk_master_csv.py --all      # 连衍生品/债券一起(默认只要股票)

跑一次就够,数据变动很慢(新股上市才变)。仓库里那份的日期写在表头。
"""
from __future__ import annotations

import argparse
import csv
import io
import sys
from pathlib import Path

HKEX_URL = ("https://www.hkex.com.hk/eng/services/trading/securities/"
            "securitieslists/ListOfSecurities.xlsx")

# 默认收哪几类。港交所全表 1.7 万条,其中衍生权证 7259 + 牛熊证 5926 ——
# 那些是**当天发行当天到期**的结构化产品,不是用户会拿来分析的标的,
# 收进来只会让 CSV 大十倍、查名字时噪音一堆。
#
# 收这三类:
#   Equity                        2803  普通股
#   Exchange Traded Products       412  ETF —— 2800/3033 这些用户天天看
#   Real Estate Investment Trusts   11  REITs —— 领展这类
_KEEP = {"Equity", "Exchange Traded Products", "Real Estate Investment Trusts"}

REPO = Path(__file__).resolve().parents[1]
OUT = REPO / "data" / "hk_master.csv"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--all", action="store_true",
                    help="连衍生权证/牛熊证/债券也收(默认只收可分析的三类)")
    ap.add_argument("--out", default=str(OUT))
    args = ap.parse_args()

    try:
        import httpx
        import openpyxl
    except ImportError as e:
        print(f"缺依赖:{e}. 需要 httpx 与 openpyxl", file=sys.stderr)
        return 1

    print(f"拉取 {HKEX_URL} …")
    try:
        r = httpx.get(HKEX_URL, headers={"User-Agent": "Mozilla/5.0"},
                      timeout=120, follow_redirects=True)
        r.raise_for_status()
    except Exception as e:                                     # noqa: BLE001
        print(f"下载失败:{type(e).__name__}: {e}", file=sys.stderr)
        return 1
    print(f"  {len(r.content) // 1024} KB")

    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # ⚠️ **必须 reset_dimensions()。**read_only 模式相信 xlsx 里那条
    # `<dimension>` 记录来判断有多少行,而港交所这个文件里那条记录是错的 ——
    # 实测只迭代出 5 行就停了(全表 2800+ 条)。
    #
    # 而它**不报错**:脚本照常写出一个 5 行的 CSV 并打印"写入成功"。
    # 这正是最难发现的一类失败 —— 除非你知道港股大概有多少只,
    # 否则看不出 5 条是错的。下面的行数下限检查就是为这个加的。
    ws.reset_dimensions()

    # 表头不在第 1 行(前两行是标题和更新日期)。**按内容找表头行,
    # 不按行号写死** —— 港交所偶尔会在上面加一行公告,写死行号的表现是
    # 某天突然导出一张全是 None 的表,而脚本照样 exit 0
    rows = ws.iter_rows(values_only=True)
    header, updated = None, ""
    for row in rows:
        cells = [str(c).strip() if c is not None else "" for c in row]
        if cells and cells[0].startswith("Updated as at"):
            updated = cells[0].replace("Updated as at", "").strip()
        if cells and cells[0] == "Stock Code":
            header = cells
            break
    if header is None:
        print("找不到表头行(第一列 'Stock Code')—— 港交所可能改了表结构。"
              "不猜着解析,请人工看一眼再改脚本", file=sys.stderr)
        return 1

    idx = {name: i for i, name in enumerate(header)}
    need = ["Stock Code", "Name of Securities", "Category", "Board Lot"]
    missing = [n for n in need if n not in idx]
    if missing:
        print(f"表头缺列 {missing} —— 结构变了,不猜着解析", file=sys.stderr)
        return 1

    out_rows, skipped = [], 0
    for row in rows:
        def g(name: str) -> str:
            i = idx[name]
            v = row[i] if i < len(row) else None
            return str(v).strip() if v is not None else ""

        code = g("Stock Code")
        if not code or not code.isdigit():
            continue
        category = g("Category")
        if not args.all and category not in _KEEP:
            skipped += 1
            continue
        # Board Lot 带千分位逗号("1,000")· 去掉再存,免得下游还要处理一次
        lot = g("Board Lot").replace(",", "")
        out_rows.append({
            "code": code.zfill(5),
            "name_en": g("Name of Securities"),
            "category": category,
            "lot_size": lot if lot.isdigit() else "",
        })

    # 行数下限 —— 港股主板+创业板长期在 2500 只以上。
    # 明显偏少说明解析出了问题(比如上面那个 dimension 坑),
    # **宁可失败也不要写一份不全的表覆盖掉好的那份** ——
    # 不全的表的表现是"某些股票查不到名字",而那看起来像那只票不存在。
    MIN_ROWS = 1500
    if len(out_rows) < MIN_ROWS:
        print(f"只解析出 {len(out_rows)} 条,少于下限 {MIN_ROWS} —— "
              f"多半是解析出了问题,不写文件。请人工看一眼 xlsx 结构",
              file=sys.stderr)
        return 1

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        f.write(f"# 港股主表 · 来自港交所官方 ListOfSecurities · 更新至 {updated}\n")
        f.write(f"# 由 scripts/gen_hk_master_csv.py 生成 · 共 {len(out_rows)} 条"
                f"{'(股票/ETF/REITs)' if not args.all else '(全部)'}\n")
        f.write("# name_en 是港交所给的英文名 —— 公开数据里没有中文名,不臆造\n")
        w = csv.DictWriter(f, fieldnames=["code", "name_en", "category", "lot_size"])
        w.writeheader()
        w.writerows(out_rows)

    print(f"写入 {out_path} · {len(out_rows)} 条"
          + (f" · 跳过非股票 {skipped} 条" if skipped else ""))
    return 0


if __name__ == "__main__":
    sys.exit(main())
